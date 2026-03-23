# -*- coding: utf-8 -*-
"""
Prompt 2 - Export demo users to a formatted Excel spreadsheet.

Run:  python backend/database/export_users_xlsx.py
Requires: pip install openpyxl
"""

import sqlite3
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH   = os.path.join(os.path.dirname(__file__), "campus_booking.db")
JSON_PATH = os.path.join(os.path.dirname(__file__), "demo_passwords.json")
OUT_PATH  = os.path.join(os.path.dirname(__file__), "JCU_Demo_User_Database.xlsx")

NAVY       = "1B3A5C"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
RED_TEXT   = "FF0000"
WARN_FILL  = "FFF3CD"

PROGRAMS = [
    "Bachelor of Information Technology",
    "Bachelor of Business",
    "Bachelor of Psychology",
    "Bachelor of Environmental Science",
    "Bachelor of Aquaculture Science",
    "Master of Information Technology",
    "Master of Business Administration",
    "Master of Data Science",
    "Diploma in Information Technology",
    "Bachelor of Education",
    "Bachelor of Hospitality Management",
    "Bachelor of Commerce",
]

# ── Style helpers ────────────────────────────────────────────────────────────

def _fill(color):
    return PatternFill("solid", fgColor=color)

def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def _style_header_row(ws, row_num, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill      = _fill(NAVY)
        cell.font      = Font(name="Arial", size=11, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _thin_border()

def _style_data_rows(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        use_gray = (r - start_row) % 2 == 1
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border    = _thin_border()
            cell.alignment = Alignment(vertical="center")
            if use_gray:
                cell.fill = _fill(LIGHT_GRAY)


# ── Sheet 1: Student Accounts ────────────────────────────────────────────────

def _build_students_sheet(wb, students, passwords):
    ws = wb.create_sheet("Student Accounts")

    ws.merge_cells("A1:I1")
    ws["A1"].value     = "JCU Campus Booking System - Demo Student Accounts"
    ws["A1"].font      = Font(name="Arial", size=14, bold=True, color=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    ws["A2"].value     = "FOR DEMO PURPOSES ONLY - Contains no real personal information"
    ws["A2"].font      = Font(name="Arial", size=10, italic=True, color=RED_TEXT)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    headers = ["#", "Student ID", "Full Name", "Email", "Password",
               "Program", "Year of Study", "Status", "Created At"]
    for col, h in enumerate(headers, 1):
        ws.cell(3, col).value = h
    _style_header_row(ws, 3, len(headers))

    for idx, row in enumerate(students, 1):
        r   = idx + 3
        sid = row["student_id"]
        ws.cell(r, 1).value = idx
        ws.cell(r, 2).value = sid
        ws.cell(r, 3).value = row["full_name"] or ""
        ws.cell(r, 4).value = row["email"] or ""
        ws.cell(r, 5).value = passwords.get(sid, "")
        ws.cell(r, 6).value = row["program"] or ""
        ws.cell(r, 7).value = row["year_of_study"] or ""
        ws.cell(r, 8).value = row["status"]
        ws.cell(r, 9).value = row["created_at"] or ""
    _style_data_rows(ws, 4, 3 + len(students), len(headers))

    ws.auto_filter.ref = f"A3:I{3 + len(students)}"
    ws.freeze_panes    = "B4"
    _set_col_widths(ws, {1: 5, 2: 14, 3: 22, 4: 28, 5: 14, 6: 36, 7: 14, 8: 14, 9: 20})


# ── Sheet 2: Staff & Admin Accounts ─────────────────────────────────────────

def _build_staff_sheet(wb, staff, passwords):
    ws = wb.create_sheet("Staff & Admin Accounts")

    ws.merge_cells("A1:H1")
    ws["A1"].value     = "JCU Campus Booking System - Demo Staff & Admin Accounts"
    ws["A1"].font      = Font(name="Arial", size=14, bold=True, color=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    ws["A2"].value     = "FOR DEMO PURPOSES ONLY - Contains no real personal information"
    ws["A2"].font      = Font(name="Arial", size=10, italic=True, color=RED_TEXT)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    headers = ["#", "Staff/Admin ID", "Full Name", "Email", "Password",
               "Role", "Status", "Created At"]
    for col, h in enumerate(headers, 1):
        ws.cell(3, col).value = h
    _style_header_row(ws, 3, len(headers))

    for idx, row in enumerate(staff, 1):
        r   = idx + 3
        sid = row["student_id"]
        ws.cell(r, 1).value = idx
        ws.cell(r, 2).value = sid
        ws.cell(r, 3).value = row["full_name"] or ""
        ws.cell(r, 4).value = row["email"] or ""
        ws.cell(r, 5).value = passwords.get(sid, "")
        ws.cell(r, 6).value = row["role"]
        ws.cell(r, 7).value = row["status"]
        ws.cell(r, 8).value = row["created_at"] or ""
    _style_data_rows(ws, 4, 3 + len(staff), len(headers))

    ws.auto_filter.ref = f"A3:H{3 + len(staff)}"
    ws.freeze_panes    = "B4"
    _set_col_widths(ws, {1: 5, 2: 16, 3: 22, 4: 28, 5: 14, 6: 10, 7: 14, 8: 20})


# ── Sheet 3: Database Summary ────────────────────────────────────────────────

def _build_summary_sheet(wb):
    ws = wb.create_sheet("Database Summary")

    ws.merge_cells("A1:D1")
    ws["A1"].value     = "Demo Database Summary"
    ws["A1"].font      = Font(name="Arial", size=14, bold=True, color=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    ws["A2"].value     = "!  DEMO DATA ONLY - NOT FOR PRODUCTION USE"
    ws["A2"].font      = Font(name="Arial", size=11, bold=True, color=RED_TEXT)
    ws["A2"].fill      = PatternFill("solid", fgColor=WARN_FILL)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    stats = [
        ("Total Students",       "=COUNTA('Student Accounts'!B4:B2404)"),
        ("Active Students",      "=COUNTIF('Student Accounts'!H4:H2404,\"active\")"),
        ("Pre-Arrival Students", "=COUNTIF('Student Accounts'!H4:H2404,\"pre_arrival\")"),
        ("Total Staff",          "=COUNTIFS('Staff & Admin Accounts'!F4:F104,\"staff\")"),
        ("Total Admins",         "=COUNTIFS('Staff & Admin Accounts'!F4:F104,\"admin\")"),
        ("Grand Total",          "=SUM(D3:D7)"),
    ]
    for offset, (label, formula) in enumerate(stats):
        r = offset + 3
        ws.cell(r, 3).value = label
        ws.cell(r, 3).font  = Font(bold=True)
        ws.cell(r, 4).value = formula
    _style_header_row(ws, 3, 4)
    for r in range(4, 9):
        ws.cell(r, 3).font = Font(bold=True)

    ws.cell(10, 3).value = "Programs Distribution"
    ws.cell(10, 3).font  = Font(bold=True, size=12)
    ws.cell(10, 4).value = "Student Count"
    _style_header_row(ws, 10, 4)

    for i, prog in enumerate(PROGRAMS):
        r = 11 + i
        ws.cell(r, 3).value = prog
        ws.cell(r, 4).value = f'=COUNTIF(\'Student Accounts\'!F4:F2404,"{prog}")'

    policy_row = 11 + len(PROGRAMS) + 2
    ws.cell(policy_row,     3).value = "Password Policy"
    ws.cell(policy_row,     3).font  = Font(bold=True, size=12)
    ws.cell(policy_row + 1, 3).value = "All passwords are 8-character alphanumeric strings"
    ws.cell(policy_row + 2, 3).value = "Passwords are bcrypt-hashed (10 rounds) in the SQLite database"
    ws.cell(policy_row + 3, 3).value = "Plaintext passwords shown here are for DEMO LOGIN ONLY"

    _set_col_widths(ws, {3: 44, 4: 16})


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    for path, label in [(DB_PATH, "SQLite DB"), (JSON_PATH, "Passwords JSON")]:
        if not os.path.exists(path):
            print(f"ERROR: {label} not found at {path}. Run generate_demo_users.py first.")
            return

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    students = conn.execute(
        "SELECT * FROM users WHERE role='student' ORDER BY student_id"
    ).fetchall()
    staff = conn.execute(
        "SELECT * FROM users WHERE role IN ('staff','admin') ORDER BY role, student_id"
    ).fetchall()
    conn.close()

    with open(JSON_PATH, encoding="utf-8") as f:
        passwords = json.load(f)

    wb = Workbook()
    wb.remove(wb.active)  # Remove default blank sheet

    _build_students_sheet(wb, students, passwords)
    _build_staff_sheet(wb, staff, passwords)
    _build_summary_sheet(wb)

    wb.save(OUT_PATH)
    print(f"Exported {len(students) + len(staff)} users to JCU_Demo_User_Database.xlsx")


if __name__ == "__main__":
    main()
